from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = Workbook()

# === SHEET 1: Overview Comparison ===
ws_overview = wb.active
ws_overview.title = "Overview"

# Headers
headers = ["Model", "Developer", "Release Date", "Context Window", "Input Price", "Output Price", "Parameters", "Key Features"]
ws_overview.append(headers)

# Data
data = [
    ["Claude 3.5 Sonnet", "Anthropic", "June 2024", "200K tokens", "$3/1M", "$15/1M", "~175B", "Vision, Tools, Multi-language"],
    ["GPT-4o", "OpenAI", "April 2024", "128K tokens", "$10/1M", "$30/1M", "Unknown", "Multimodal, Vision, Function calling"],
    ["Grok 2", "xAI", "2024", "131K tokens", "$2/1M", "$10/1M", "Unknown", "Multi-language, Tools, Text-focused"],
    ["Gemini 2.0 Flash", "Google", "Dec 2024", "1M tokens", "$0.10/1M", "$0.28/1M", "Unknown", "Multimodal, Native tools, Code execution"],
    ["DeepSeek V3", "DeepSeek", "Dec 2024", "128K tokens", "$0.07/1M", "$1.68/1M", "671B (37B active)", "MoE architecture, Open-source, Multi-token prediction"],
]

for row in data:
    ws_overview.append(row)

# Styling
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for cell in ws_overview[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

for row in ws_overview.iter_rows(min_row=2, max_row=6, min_col=1, max_col=8):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# Column widths
column_widths = [20, 15, 15, 15, 12, 12, 20, 35]
for i, width in enumerate(column_widths, 1):
    ws_overview.column_dimensions[get_column_letter(i)].width = width

# Add table with filters
tab = Table(displayName="OverviewTable", ref=f"A1:H{len(data)+1}")
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws_overview.add_table(tab)

# === SHEET 2: Detailed Specifications ===
ws_specs = wb.create_sheet("Detailed Specs")

specs_headers = ["Model", "Context Window (tokens)", "Max Output", "Input Price ($/1M)", "Output Price ($/1M)",
                 "Parameters", "Architecture", "Fine-tuning", "API Access"]
ws_specs.append(specs_headers)

specs_data = [
    ["Claude 3.5 Sonnet", 200000, 8192, 3.00, 15.00, "~175B", "Dense", "No", "Anthropic API, AWS Bedrock, GCP Vertex AI"],
    ["GPT-4o", 128000, 16384, 10.00, 30.00, "Unknown", "Dense", "Yes", "OpenAI API, Azure OpenAI"],
    ["Grok 2", 131072, 131072, 2.00, 10.00, "Unknown", "Dense", "No", "xAI API"],
    ["Gemini 2.0 Flash", 1000000, 8192, 0.10, 0.28, "Unknown", "Dense", "No", "Google AI Studio, Vertex AI"],
    ["DeepSeek V3", 128000, 8192, 0.07, 1.68, "671B (37B active)", "MoE", "No", "DeepSeek API"],
]

for row in specs_data:
    ws_specs.append(row)

# Styling
for cell in ws_specs[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

for row in ws_specs.iter_rows(min_row=2, max_row=6, min_col=1, max_col=9):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

specs_widths = [20, 18, 12, 16, 16, 20, 15, 12, 40]
for i, width in enumerate(specs_widths, 1):
    ws_specs.column_dimensions[get_column_letter(i)].width = width

# Add table
tab2 = Table(displayName="SpecsTable", ref=f"A1:I{len(specs_data)+1}")
tab2.tableStyleInfo = style
ws_specs.add_table(tab2)

# === SHEET 3: Benchmark Performance ===
ws_bench = wb.create_sheet("Benchmarks")

bench_headers = ["Model", "MMLU (%)", "HumanEval (%)", "Coding Ability", "Reasoning", "Long Context", "Multimodal"]
ws_bench.append(bench_headers)

bench_data = [
    ["Claude 3.5 Sonnet", 81.5, 88, "Excellent", "Excellent", "Very Good (200K)", "Yes (Vision)"],
    ["GPT-4o", 89, 90.2, "Excellent", "Excellent", "Very Good (128K)", "Yes (Vision, Audio)"],
    ["Grok 2 / Grok 3", 92.7, 86.5, "Very Good", "Excellent", "Good (131K)", "Limited"],
    ["Gemini 2.0 Flash", 80.5, 85, "Very Good", "Very Good", "Excellent (1M)", "Yes (Audio, Video, Images)"],
    ["DeepSeek V3", 79.5, 86, "Very Good", "Very Good", "Very Good (128K)", "No"],
]

for row in bench_data:
    ws_bench.append(row)

# Styling
for cell in ws_bench[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

for row in ws_bench.iter_rows(min_row=2, max_row=6, min_col=1, max_col=7):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

bench_widths = [20, 12, 14, 15, 15, 20, 25]
for i, width in enumerate(bench_widths, 1):
    ws_bench.column_dimensions[get_column_letter(i)].width = width

# Add table
tab3 = Table(displayName="BenchTable", ref=f"A1:G{len(bench_data)+1}")
tab3.tableStyleInfo = style
ws_bench.add_table(tab3)

# Add chart for MMLU scores
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "MMLU Benchmark Comparison"
chart.y_axis.title = "Score (%)"
chart.x_axis.title = "Model"

data_ref = Reference(ws_bench, min_col=2, min_row=1, max_row=6)
cats_ref = Reference(ws_bench, min_col=1, min_row=2, max_row=6)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)

ws_bench.add_chart(chart, "I2")

# === SHEET 4: Pricing Comparison ===
ws_price = wb.create_sheet("Pricing Analysis")

price_headers = ["Model", "Input ($/1M tokens)", "Output ($/1M tokens)", "Total for 1M in + 1M out",
                 "Cost Efficiency", "Best For"]
ws_price.append(price_headers)

price_data = [
    ["Claude 3.5 Sonnet", 3.00, 15.00, "=B2+C2", "Mid-range", "Balanced performance and cost"],
    ["GPT-4o", 10.00, 30.00, "=B3+C3", "Premium", "High-quality outputs, complex tasks"],
    ["Grok 2", 2.00, 10.00, "=B4+C4", "Good", "Cost-effective reasoning"],
    ["Gemini 2.0 Flash", 0.10, 0.28, "=B5+C5", "Excellent", "Large context, budget-conscious"],
    ["DeepSeek V3", 0.07, 1.68, "=B6+C6", "Excellent", "Cost-effective at scale"],
]

for row in price_data:
    ws_price.append(row)

# Styling
for cell in ws_price[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

for row in ws_price.iter_rows(min_row=2, max_row=6, min_col=1, max_col=6):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

# Format currency columns
for row in ws_price.iter_rows(min_row=2, max_row=6, min_col=2, max_col=4):
    for cell in row:
        cell.number_format = '$#,##0.00'

price_widths = [20, 18, 18, 20, 15, 30]
for i, width in enumerate(price_widths, 1):
    ws_price.column_dimensions[get_column_letter(i)].width = width

# Add table
tab4 = Table(displayName="PriceTable", ref=f"A1:F{len(price_data)+1}")
tab4.tableStyleInfo = style
ws_price.add_table(tab4)

# === SHEET 5: Use Case Recommendations ===
ws_usecase = wb.create_sheet("Use Cases")

usecase_headers = ["Use Case", "Best Model", "Alternative", "Reasoning"]
ws_usecase.append(usecase_headers)

usecase_data = [
    ["Code Generation", "GPT-4o", "Claude 3.5 Sonnet", "GPT-4o leads HumanEval (90.2%), Claude close second (88%)"],
    ["Long Document Analysis", "Gemini 2.0 Flash", "Claude 3.5 Sonnet", "Gemini's 1M context window vs Claude's 200K"],
    ["Cost-Effective Deployment", "DeepSeek V3", "Gemini 2.0 Flash", "DeepSeek offers lowest pricing with good performance"],
    ["Complex Reasoning", "Grok 3", "GPT-4o", "Grok 3 leads MMLU (92.7%) for reasoning tasks"],
    ["Multimodal Tasks", "Gemini 2.0 Flash", "GPT-4o", "Gemini supports audio, video, images natively"],
    ["Balanced Performance", "Claude 3.5 Sonnet", "GPT-4o", "Strong across all metrics with reasonable pricing"],
    ["Creative Writing", "Claude 3.5 Sonnet", "GPT-4o", "Known for high-quality long-form content"],
    ["Enterprise Integration", "GPT-4o", "Claude 3.5 Sonnet", "Wide API availability (Azure, OpenAI)"],
]

for row in usecase_data:
    ws_usecase.append(row)

# Styling
for cell in ws_usecase[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

for row in ws_usecase.iter_rows(min_row=2, max_row=9, min_col=1, max_col=4):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

usecase_widths = [25, 20, 20, 50]
for i, width in enumerate(usecase_widths, 1):
    ws_usecase.column_dimensions[get_column_letter(i)].width = width

# Add table
tab5 = Table(displayName="UseCaseTable", ref=f"A1:D{len(usecase_data)+1}")
tab5.tableStyleInfo = style
ws_usecase.add_table(tab5)

# === SHEET 6: Summary Dashboard ===
ws_summary = wb.create_sheet("Summary", 0)

# Title
ws_summary['A1'] = "AI Model Comparison Dashboard"
ws_summary['A1'].font = Font(size=18, bold=True, color="1F4E78")
ws_summary.merge_cells('A1:F1')
ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws_summary.row_dimensions[1].height = 30

# Key Insights
ws_summary['A3'] = "Key Insights"
ws_summary['A3'].font = Font(size=14, bold=True, color="1F4E78")

insights = [
    ["Lowest Cost:", "DeepSeek V3 ($0.07/$1.68) and Gemini 2.0 Flash ($0.10/$0.28)"],
    ["Best Reasoning:", "Grok 3 (MMLU 92.7%)"],
    ["Best Coding:", "GPT-4o (HumanEval 90.2%)"],
    ["Largest Context:", "Gemini 2.0 Flash (1M tokens)"],
    ["Best Balance:", "Claude 3.5 Sonnet (strong performance, reasonable pricing)"],
]

row_num = 4
for insight in insights:
    ws_summary[f'A{row_num}'] = insight[0]
    ws_summary[f'B{row_num}'] = insight[1]
    ws_summary[f'A{row_num}'].font = Font(bold=True)
    row_num += 1

# Quick Stats
ws_summary['A11'] = "Quick Statistics"
ws_summary['A11'].font = Font(size=14, bold=True, color="1F4E78")

stats_headers = ["Metric", "Min", "Max", "Average"]
ws_summary.append([])  # Empty row
ws_summary.append(stats_headers)

stats_data = [
    ["Context Window", "128K", "1M", "~350K"],
    ["Input Price", "$0.07", "$10.00", "$3.03"],
    ["Output Price", "$0.28", "$30.00", "$11.39"],
    ["MMLU Score", "79.5%", "92.7%", "84.6%"],
]

for row in stats_data:
    ws_summary.append(row)

for cell in ws_summary[13]:
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    cell.font = Font(bold=True)
    cell.border = thin_border

for row in ws_summary.iter_rows(min_row=14, max_row=17, min_col=1, max_col=4):
    for cell in row:
        cell.border = thin_border

ws_summary.column_dimensions['A'].width = 20
ws_summary.column_dimensions['B'].width = 40
ws_summary.column_dimensions['C'].width = 15
ws_summary.column_dimensions['D'].width = 15

# Instructions
ws_summary['A20'] = "How to Use This Workbook"
ws_summary['A20'].font = Font(size=14, bold=True, color="1F4E78")

instructions = [
    "• Overview: Quick comparison of all models",
    "• Detailed Specs: Technical specifications and API access",
    "• Benchmarks: Performance metrics (MMLU, HumanEval)",
    "• Pricing Analysis: Cost comparison and efficiency",
    "• Use Cases: Recommended models for specific tasks",
    "• All tables have filters enabled - click dropdown arrows in headers",
]

row_num = 21
for instruction in instructions:
    ws_summary[f'A{row_num}'] = instruction
    row_num += 1

# Save
output_path = "C:\\Users\\LEEDONGGEUN\\.claude\\skills\\xlsx\\workspace\\AI_Models_Comparison_2025.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
